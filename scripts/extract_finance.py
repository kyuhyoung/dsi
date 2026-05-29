"""재무제표 xlsx → kb/company/<name>/finance.yaml 자동 변환.

일반화:
  - 회사·시트 이름·행 위치 가정 *없음*.
  - templates/finance_label_map.yaml 의 keywords 로 *계정과목 셀* 매칭.
  - header 행에서 FY20XX 패턴으로 *연도 컬럼* 자동 식별.
  - 시트 어디서든 값 찾으면 records[year][field] 에 채움.
  - 미매칭 필드는 yaml 에 안 적음 (빌더가 '확인 필요' 표기 담당).

용법:
    python scripts/extract_finance.py <input.xlsx> <output.yaml> [--company "회사명"]

설계 결정:
  - *합계 행 우선*: '매출액 합계' 같은 합계 행이 항목 행보다 점수 높음.
  - *덜 들여쓴 행 우선*: 같은 키워드 중 들여쓰기 적은 (= 합계/총계 row) 우선.
  - 값 형식: 원 단위 정수 (xlsx 의 원 단위 그대로 보존, 표준 스키마와 일치).
"""
import sys
import re
from pathlib import Path
import yaml
import openpyxl


# 연도 헤더 패턴 — 'FY2024' '2024' '2024년' 등 일반 흡수
YEAR_HEADER_RE = re.compile(r"(?:FY|fy|F/Y\s*)?(\d{4})\s*(?:년|FY|fy)?$")

# 합계/총계 키워드 — yaml 정본 (finance_label_map.yaml.total_row_keywords)에서 로드.
# 빈 list = fallback (yaml 로드 실패 시 0개). 안전 default 없음 — yaml 누락이 더 큰 문제.
TOTAL_KEYWORDS = []


def load_label_map(project_root: Path) -> dict:
    """templates/finance_label_map.yaml 로드 + 모듈 전역 TOTAL_KEYWORDS 갱신."""
    global TOTAL_KEYWORDS
    p = project_root / "templates" / "finance_label_map.yaml"
    data = yaml.safe_load(p.read_text(encoding="utf-8"))
    TOTAL_KEYWORDS = list(data.get("total_row_keywords") or [])
    return data


def normalize_label(text: str) -> str:
    """라벨 정규화: 들여쓰기·괄호·기호 제거 + 소문자."""
    if not text:
        return ""
    s = text.strip()
    # 로마숫자 + 점 + 공백 제거 ('Ⅰ. ', 'Ⅱ.' 등)
    s = re.sub(r"^[ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ]+\.\s*", "", s)
    # 괄호와 안 내용 제거 ('(GPM%)' 등)
    s = re.sub(r"\([^)]*\)", "", s)
    return s.strip()


def is_indented(text: str) -> int:
    """라벨 들여쓰기 정도. 0=합계/대분류, 양수=세부 항목."""
    if not text:
        return 0
    return len(text) - len(text.lstrip())


def find_year_columns(ws) -> dict:
    """header 행 검색 → 연도 컬럼 → 'YYYY' 매핑 반환."""
    # header 가 어디 있는지 모름 — 첫 30행 안에서 *FY20XX 가 가장 많은 행* 을 header 로 추정
    best_row, best_map = 0, {}
    for r in range(1, min(ws.max_row + 1, 30)):
        col_to_year = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            m = YEAR_HEADER_RE.match(v.strip())
            if m:
                col_to_year[c] = m.group(1)
        if len(col_to_year) > len(best_map):
            best_map = col_to_year
            best_row = r
    return best_map  # {col_idx: 'YYYY'}


def score_row_match(label: str, keyword: str) -> int:
    """라벨이 keyword 를 포함하는지. 합계 행이면 보너스, 들여쓰기 페널티."""
    norm = normalize_label(label)
    if keyword not in norm:
        return -1  # 매칭 실패
    score = 100
    # 합계 행 보너스
    if any(t in norm for t in TOTAL_KEYWORDS):
        score += 50
    # 들여쓰기 페널티
    score -= is_indented(label) * 2
    # 정확 일치 보너스 (keyword 가 라벨 전체)
    if norm.strip() == keyword:
        score += 30
    return score


def extract_field_records(wb, field_name: str, keywords: list) -> dict:
    """field 의 keyword 들로 모든 시트 검색 → {year: value} 반환."""
    best_score = -1
    best_records = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        year_cols = find_year_columns(ws)
        if not year_cols:
            continue
        # 모든 행 + 모든 텍스트 셀에서 keyword 매칭
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if not isinstance(v, str):
                    continue
                # 각 keyword 시도, 가장 높은 score 선택
                row_score = -1
                for kw in keywords:
                    s = score_row_match(v, kw)
                    if s > row_score:
                        row_score = s
                if row_score < 0:
                    continue
                # 이 행의 연도별 값 수집
                records = {}
                for col, year in year_cols.items():
                    cell_v = ws.cell(r, col).value
                    if isinstance(cell_v, (int, float)) and not isinstance(cell_v, bool):
                        records[year] = int(cell_v)
                # 행에 값이 있으면 후보
                if records and row_score > best_score:
                    best_score = row_score
                    best_records = records
    return best_records


def extract_meta(wb, label_map: dict) -> dict:
    """표지·헤더에서 회사명·단위 등 메타 추출.
    키워드·단위 사전 모두 yaml 정본 (finance_label_map.yaml.meta_labels + .unit_patterns).
    코드에 표기 박지 않음 — 새 표기 추가는 yaml 만.
    """
    meta_labels = label_map.get("meta_labels") or {}
    company_keywords = list(meta_labels.get("company_name") or [])
    unit_keywords = list(meta_labels.get("currency_unit") or [])
    # 단위 alternation도 yaml unit_patterns 에서
    unit_patterns = label_map.get("unit_patterns") or []
    unit_names = [str(p["pattern"]) for p in unit_patterns if p.get("pattern")]
    # 추가 통화 키 (KRW, USD 등) — yaml 의 normalize.strip_units 에서도 가져옴
    extra_currencies = [u for u in (label_map.get("normalize") or {}).get("strip_units") or []
                        if u not in unit_names]
    all_unit_names = sorted(set(unit_names + extra_currencies), key=len, reverse=True)
    unit_alt_re = re.compile("|".join(re.escape(u) for u in all_unit_names)) if all_unit_names else None

    meta = {"currency": "KRW"}
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in range(1, min(ws.max_row + 1, 15)):
            for c in range(1, min(ws.max_column + 1, 8)):
                v = ws.cell(r, c).value
                if not isinstance(v, str):
                    continue
                v_str = v.strip()
                # 회사명 키워드 매칭
                if any(kw in v_str for kw in company_keywords):
                    for cc in range(c + 1, min(c + 3, ws.max_column + 1)):
                        nv = ws.cell(r, cc).value
                        if isinstance(nv, str) and nv.strip():
                            meta["company_name"] = nv.strip()
                            break
                # 단위 키워드 매칭
                if any(kw in v_str for kw in unit_keywords):
                    for cc in range(c + 1, min(c + 3, ws.max_column + 1)):
                        nv = ws.cell(r, cc).value
                        if isinstance(nv, str) and nv.strip() and unit_alt_re is not None:
                            m = unit_alt_re.search(nv)
                            if m:
                                meta["currency_unit"] = m.group(0)
                            break
        if "company_name" in meta:
            break
    return meta


def extract_finance(xlsx_path: Path, label_map: dict, project_root: Path) -> dict:
    """xlsx → finance.yaml dict."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # 필드별 records 추출
    records_by_year = {}
    fields = label_map.get("fields", {})
    for field_name, field_def in fields.items():
        keywords = field_def.get("keywords") or []
        if not keywords:
            continue
        records = extract_field_records(wb, field_name, keywords)
        for year, value in records.items():
            records_by_year.setdefault(year, {})[field_name] = value

    # 메타 (키워드 yaml 정본)
    meta = extract_meta(wb, label_map)
    meta["source"] = str(xlsx_path.name)

    # 연도 키 내림차순 정렬 (가장 최근 위)
    sorted_records = {y: records_by_year[y] for y in sorted(records_by_year.keys(), reverse=True)}

    return {
        "meta": meta,
        "records": sorted_records,
    }


def main():
    if len(sys.argv) < 3:
        print("사용: python scripts/extract_finance.py <input.xlsx> <output.yaml> [--company \"회사명\"]", file=sys.stderr)
        sys.exit(1)
    inp = Path(sys.argv[1])
    outp = Path(sys.argv[2])
    project_root = Path(__file__).parent.parent

    label_map = load_label_map(project_root)
    result = extract_finance(inp, label_map, project_root)

    # --company override
    if "--company" in sys.argv:
        i = sys.argv.index("--company")
        if i + 1 < len(sys.argv):
            result["meta"]["company_name"] = sys.argv[i + 1]

    outp.parent.mkdir(parents=True, exist_ok=True)
    outp.write_text(
        yaml.dump(result, allow_unicode=True, sort_keys=False, default_flow_style=False),
        encoding="utf-8",
    )
    print(f"[OK] {inp} -> {outp}", file=sys.stderr)
    print(f"  연도: {list(result['records'].keys())}", file=sys.stderr)
    for y, rec in result["records"].items():
        print(f"  {y}: {len(rec)} fields", file=sys.stderr)


if __name__ == "__main__":
    main()
